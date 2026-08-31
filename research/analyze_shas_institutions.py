# /// script
# dependencies = [
#   "numpy>=2.0",
#   "pandas>=2.2",
#   "requests>=2.32",
#   "rapidfuzz>=3.9",
#   "scikit-learn>=1.5",
#   "openpyxl>=3.1"
# ]
# ///

from __future__ import annotations

import json
import math
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz
from sklearn.neighbors import BallTree

OUT = Path("research/final_output")
OUT.mkdir(parents=True, exist_ok=True)
S = requests.Session()
S.headers.update({"User-Agent": "shas-2022-institution-research/1.0", "Accept": "*/*"})

BALLOT_RESOURCE = "cc223336-07bc-485d-b160-62df92967c0a"
MOSDOT_RESOURCE = "5548fd63-5868-4053-ad81-98caddc5e232"
COORDS_RESOURCE = "5c5d6bb0-755d-470d-84b6-d7dd3135ba9c"
LOCATIONS_URL = "https://raw.githubusercontent.com/harelc/elections-vote-transfer/master/data/ballot_locations_25.json"

HAREDI_LOCALITIES = {
    "בני ברק",
    "אלעד",
    "מודיעין עילית",
    "ביתר עילית",
    "עמנואל",
    "רכסים",
    "קרית יערים",
    "קריית יערים",
}

HEBREW_DIACRITICS = re.compile(r"[\u0591-\u05C7]")
NON_ALNUM = re.compile(r"[^0-9A-Za-z\u0590-\u05FF]+")
SPACES = re.compile(r"\s+")


def fetch_datastore(resource_id: str, page_size: int = 10000) -> tuple[pd.DataFrame, dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    fields: list[dict[str, Any]] = []
    total: int | None = None
    while True:
        params = {
            "resource_id": resource_id,
            "limit": page_size,
            "offset": offset,
            "include_total": "true",
        }
        r = S.get("https://data.gov.il/api/3/action/datastore_search", params=params, timeout=180)
        r.raise_for_status()
        payload = r.json()
        if not payload.get("success"):
            raise RuntimeError(payload)
        result = payload["result"]
        if total is None:
            total = int(result.get("total") or 0)
            fields = result.get("fields", [])
        batch = result.get("records", [])
        rows.extend(batch)
        if not batch or len(rows) >= (total or 0) or len(batch) < page_size:
            break
        offset += len(batch)
    return pd.DataFrame(rows), {"total": total, "rows_fetched": len(rows), "fields": fields}


def get_json(url: str) -> Any:
    r = S.get(url, timeout=180)
    r.raise_for_status()
    return r.json()


def clean_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    s = unicodedata.normalize("NFKC", str(value))
    s = HEBREW_DIACRITICS.sub("", s)
    s = s.replace("״", '"').replace("׳", "'")
    s = s.replace("–", "-").replace("—", "-")
    s = NON_ALNUM.sub(" ", s)
    return SPACES.sub(" ", s).strip().lower()


def norm_city(value: Any) -> str:
    s = clean_text(value)
    replacements = {
        "תל אביב יפו": "תל אביב יפו",
        "קרית יערים טלז סטון": "קרית יערים",
        "קריית יערים טלז סטון": "קרית יערים",
        "מודיעין מכבים רעות": "מודיעין מכבים רעות",
        "נצרת עילית": "נוף הגליל",
        "קרית מלאכי": "קרית מלאכי",
        "קריית מלאכי": "קרית מלאכי",
        "קרית גת": "קרית גת",
        "קריית גת": "קרית גת",
        "קרית אתא": "קרית אתא",
        "קריית אתא": "קרית אתא",
        "קרית ביאליק": "קרית ביאליק",
        "קריית ביאליק": "קרית ביאליק",
        "קרית אונו": "קרית אונו",
        "קריית אונו": "קרית אונו",
        "קרית שמונה": "קרית שמונה",
        "קריית שמונה": "קרית שמונה",
        "קרית ים": "קרית ים",
        "קריית ים": "קרית ים",
        "קרית מוצקין": "קרית מוצקין",
        "קריית מוצקין": "קרית מוצקין",
    }
    return replacements.get(s, s)


def norm_school(value: Any) -> str:
    s = clean_text(value)
    repl = {
        "ביה ס": "בית ספר",
        "בי ס": "בית ספר",
        "ביס": "בית ספר",
        "ביהס": "בית ספר",
        "ת ת": "תלמוד תורה",
        "תת": "תלמוד תורה",
        "ממ ח": "ממלכתי חרדי",
        "ממח": "ממלכתי חרדי",
        "ממ ד": "ממלכתי דתי",
        "ממד": "ממלכתי דתי",
        "חט ב": "חטיבת ביניים",
        "חטב": "חטיבת ביניים",
    }
    for a, b in repl.items():
        s = re.sub(rf"\b{re.escape(a)}\b", b, s)
    generic_phrases = [
        "בית הספר",
        "בית ספר",
        "בית חינוך",
        "קרית חינוך",
        "קריית חינוך",
        "מרכז חינוכי",
        "מרכז חינוך",
        "מוסדות חינוך",
        "מוסד חינוכי",
    ]
    for phrase in generic_phrases:
        s = s.replace(phrase, " ")
    generic_tokens = {
        "מבנה", "אגף", "חדר", "קומה", "כניסה", "ימני", "שמאלי", "גדול", "קטן",
        "חדש", "חדשה", "ישן", "ישנה", "עירוני", "עירונית", "מתחם", "שלוחה",
    }
    tokens = [t for t in s.split() if t not in generic_tokens]
    return " ".join(tokens).strip()


def explicit_haredi_marker(value: Any) -> bool:
    s = clean_text(value)
    markers = [
        "בית יעקב", "תלמוד תורה", "חינוך עצמאי", "בני יוסף", "מעיין החינוך",
        "מעין החינוך", "אל המעיין", "ממלכתי חרדי", "ממ ח", "ממח",
        "אגודת ישראל", "נשי אגודת ישראל", "סמינר בית יעקב",
    ]
    return any(m in s for m in markers)


def network_tag(row: pd.Series) -> str:
    s = " ".join(clean_text(row.get(c, "")) for c in ["שם מוסד", "גורם מדווח", "מעמד משפטי"])
    if any(x in s for x in ["בני יוסף", "מעיין החינוך", "מעין החינוך", "אל המעיין"]):
        return "בני יוסף/מעיין החינוך"
    if "חינוך עצמאי" in s:
        return "החינוך העצמאי"
    if any(x in s for x in ["ממלכתי חרדי", "ממ ח", "ממח"]):
        return "ממלכתי חרדי"
    return ""


def ballot_str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    try:
        x = float(s)
        if x.is_integer():
            return str(int(x))
        return (f"{x:.6f}").rstrip("0").rstrip(".")
    except Exception:
        return s


def safe_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").fillna(0)


def match_score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if a == b:
        return 100.0
    short, long = (a, b) if len(a) <= len(b) else (b, a)
    contain = 0.0
    if len(short) >= 4 and short in long:
        contain = 95.0 + min(4.0, 4.0 * len(short) / max(1, len(long)))
    scores = [
        float(fuzz.WRatio(a, b)),
        float(fuzz.token_set_ratio(a, b)),
        max(0.0, float(fuzz.partial_ratio(a, b)) - 3.0),
        contain,
    ]
    result = max(scores)
    if min(len(a), len(b)) <= 3:
        result -= 15
    return max(0.0, min(100.0, result))


def institution_type_weight(row: pd.Series) -> float:
    txt = " ".join(clean_text(row.get(c, "")) for c in ["סוג מוסד", "סוג מסגרת אירגונית", "שם מוסד", "משכבה", "עד שכבה"])
    if "גן ילדים" in txt or re.search(r"\bגן\b", txt):
        return 0.40
    if any(x in txt for x in ["יסודי", "תלמוד תורה"]):
        return 1.00
    if any(x in txt for x in ["חטיבת ביניים", "חטיבה"]):
        return 0.70
    if any(x in txt for x in ["ישיבה", "סמינר", "על יסודי", "תיכון"]):
        return 0.45
    return 0.65


def add_sheet(wb: Workbook, title: str, frame: pd.DataFrame, widths: dict[str, int] | None = None) -> None:
    ws = wb.create_sheet(title=title[:31])
    ws.sheet_view.rightToLeft = True
    if frame.empty:
        ws.append(["אין נתונים"])
        return
    cols = [str(c) for c in frame.columns]
    ws.append(cols)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    for row in frame.itertuples(index=False, name=None):
        ws.append([None if (isinstance(v, float) and math.isnan(v)) else v for v in row])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right", vertical="top", wrap_text=False)
    for idx, col in enumerate(cols, 1):
        if widths and col in widths:
            width = widths[col]
        else:
            sample = [len(str(col))]
            for v in frame[col].head(200).astype(str):
                sample.append(len(v))
            width = min(45, max(10, int(np.percentile(sample, 90)) + 2))
        ws.column_dimensions[get_column_letter(idx)].width = width


def main() -> None:
    audit: dict[str, Any] = {}

    ballots, ballots_meta = fetch_datastore(BALLOT_RESOURCE)
    mos_all, mos_meta = fetch_datastore(MOSDOT_RESOURCE)
    coords, coords_meta = fetch_datastore(COORDS_RESOURCE)
    locations = get_json(LOCATIONS_URL)
    ballot_to_location: dict[str, str] = locations.get("ballot_to_location", {})

    audit["sources"] = {
        "ballots": ballots_meta,
        "institutions": mos_meta,
        "coordinates": coords_meta,
        "location_keys": len(ballot_to_location),
    }

    # Normalize official election data.
    for c in ["סמל ישוב", "כשרים", "שס", "ג", "מחל", "ט", "פה", "כן", "אמת", "מרצ", "ל"]:
        if c in ballots:
            ballots[c] = safe_num(ballots[c])
    ballots["קלפי_טקסט"] = ballots["קלפי"].map(ballot_str)
    ballots["מפתח_קלפי"] = ballots["סמל ישוב"].astype(int).astype(str) + ":" + ballots["קלפי_טקסט"]
    ballots["מיקום קלפי"] = ballots["מפתח_קלפי"].map(ballot_to_location).fillna("")
    ballots["עיר_נורמלית"] = ballots["שם ישוב"].map(norm_city)
    ballots["חיצוניות"] = (ballots["סמל ישוב"] == 9999) | ballots["שם ישוב"].astype(str).str.contains("מעטפות", na=False)
    ballots["שיעור ג"] = np.where(ballots["כשרים"] > 0, ballots["ג"] / ballots["כשרים"], 0.0)
    ballots["שיעור שס"] = np.where(ballots["כשרים"] > 0, ballots["שס"] / ballots["כשרים"], 0.0)
    ballots["שיעור חרדי משולב"] = np.where(ballots["כשרים"] > 0, (ballots["ג"] + ballots["שס"]) / ballots["כשרים"], 0.0)
    ballots["יישוב חרדי מובהק"] = ballots["עיר_נורמלית"].isin({norm_city(x) for x in HAREDI_LOCALITIES})

    # Select the institution snapshot for 2022, locally rather than via CKAN filter.
    mos_all["שנה_מספר"] = pd.to_numeric(mos_all.get("שנה"), errors="coerce")
    year_counts = mos_all["שנה_מספר"].value_counts(dropna=True).sort_index()
    available_years = sorted(int(y) for y in year_counts.index if pd.notna(y))
    if 2022 in available_years:
        target_year = 2022
    else:
        past = [y for y in available_years if y <= 2022]
        target_year = max(past) if past else min(available_years)
    mos = mos_all[mos_all["שנה_מספר"] == target_year].copy()
    mos["סמל_טקסט"] = mos["סמל מוסד"].map(ballot_str)
    mos["עיר_מוסד"] = mos["שם ישוב"].map(norm_city)
    mos["רשות_מוסד"] = mos["שם רשות"].map(norm_city)
    mos["שם_מוסד_נורמלי"] = mos["שם מוסד"].map(norm_school)
    mos["פיקוח_נורמלי"] = mos["פיקוח"].map(clean_text)
    mos["חרדי"] = mos["פיקוח_נורמלי"].str.contains("חרדי", na=False) | mos["שם מוסד"].map(explicit_haredi_marker)
    mos["רשת_מזוהה"] = mos.apply(network_tag, axis=1)
    mos["תלמידים"] = pd.to_numeric(mos["סהכ תלמידים במוסד"], errors="coerce")
    fallback_students = np.where(mos["סוג מוסד"].astype(str).str.contains("גן", na=False), 28.0, 120.0)
    mos["תלמידים_מושלמים"] = mos["תלמידים"].where(mos["תלמידים"].notna() & (mos["תלמידים"] > 0), fallback_students)
    mos["משקל_סוג"] = mos.apply(institution_type_weight, axis=1)
    mos["תלמידים_שקולים"] = mos["תלמידים_מושלמים"].clip(upper=1200) * mos["משקל_סוג"]

    coords["סמל_טקסט"] = coords["SEMEL_MOSAD"].map(ballot_str)
    coords["lon"] = pd.to_numeric(coords["UTM_X"], errors="coerce")
    coords["lat"] = pd.to_numeric(coords["UTM_Y"], errors="coerce")
    coords["דיוק"] = coords["RAMAT_DIYUK_MIKUM"].map(clean_text)
    coords = coords.drop_duplicates("סמל_טקסט", keep="first")
    mos = mos.merge(coords[["סמל_טקסט", "lon", "lat", "דיוק"]], on="סמל_טקסט", how="left")
    mos["קואורדינטה_תקינה"] = mos["lat"].between(29.0, 34.0) & mos["lon"].between(34.0, 36.0)
    mos["קואורדינטה_לצפיפות"] = mos["קואורדינטה_תקינה"] & ~mos["דיוק"].str.contains("נמוכה", na=False)

    audit["institutions"] = {
        "available_years": available_years,
        "year_counts": {str(int(k)): int(v) for k, v in year_counts.items()},
        "selected_year": target_year,
        "selected_rows": int(len(mos)),
        "haredi_rows": int(mos["חרדי"].sum()),
        "with_good_coordinates": int(mos["קואורדינטה_לצפיפות"].sum()),
        "coordinate_accuracy": coords["דיוק"].value_counts(dropna=False).to_dict(),
        "explicit_network_tags": mos["רשת_מזוהה"].value_counts().to_dict(),
    }

    # Build site-level table; several ballot boxes can share one site.
    regular = ballots[~ballots["חיצוניות"]].copy()
    sites = regular[["סמל ישוב", "שם ישוב", "עיר_נורמלית", "מיקום קלפי"]].drop_duplicates().copy()
    sites["מיקום_נורמלי"] = sites["מיקום קלפי"].map(norm_school)
    sites["סמן_חרדי_בשם"] = sites["מיקום קלפי"].map(explicit_haredi_marker)
    sites["מזהה_אתר"] = sites["סמל ישוב"].astype(int).astype(str) + "|" + sites["מיקום קלפי"].astype(str)

    city_index: dict[str, list[int]] = defaultdict(list)
    for idx, row in mos.iterrows():
        for city_key in {row["עיר_מוסד"], row["רשות_מוסד"]}:
            if city_key:
                city_index[city_key].append(idx)

    match_rows: list[dict[str, Any]] = []
    for _, site in sites.iterrows():
        candidate_indices = list(dict.fromkeys(city_index.get(site["עיר_נורמלית"], [])))
        query = site["מיקום_נורמלי"]
        ranked: list[tuple[float, int]] = []
        if query and candidate_indices:
            for idx in candidate_indices:
                cand = mos.at[idx, "שם_מוסד_נורמלי"]
                sc = match_score(query, cand)
                if sc >= 45:
                    ranked.append((sc, idx))
        ranked.sort(reverse=True, key=lambda x: x[0])
        best_score = ranked[0][0] if ranked else 0.0
        second_score = ranked[1][0] if len(ranked) > 1 else 0.0
        best_idx = ranked[0][1] if ranked else None
        top_indices = [idx for sc, idx in ranked[:3] if sc >= best_score - 3]
        top_haredi = {bool(mos.at[idx, "חרדי"]) for idx in top_indices}
        same_class = len(top_haredi) == 1 and bool(top_indices)
        accepted = bool(best_idx is not None and best_score >= 82 and (best_score >= 94 or best_score - second_score >= 4 or same_class))
        anchor = bool(
            accepted
            and best_idx is not None
            and best_score >= 90
            and bool(mos.at[best_idx, "קואורדינטה_לצפיפות"])
            and (best_score >= 96 or best_score - second_score >= 5)
        )
        best = mos.loc[best_idx] if best_idx is not None else None
        matched_haredi = bool(accepted and best is not None and best["חרדי"])
        # Explicit markers can classify a site when the top ambiguous candidates are consistently Haredi.
        if site["סמן_חרדי_בשם"] and same_class and top_haredi == {True} and best_score >= 70:
            matched_haredi = True
        match_rows.append({
            "מזהה_אתר": site["מזהה_אתר"],
            "ציון התאמה": round(best_score, 2),
            "פער התאמה": round(best_score - second_score, 2),
            "התאמה התקבלה": accepted,
            "עוגן קואורדינטות": anchor,
            "סמל מוסד מותאם": best["סמל_טקסט"] if best is not None else "",
            "שם מוסד מותאם": best["שם מוסד"] if best is not None else "",
            "פיקוח מוסד מותאם": best["פיקוח"] if best is not None else "",
            "סוג מוסד מותאם": best["סוג מוסד"] if best is not None else "",
            "רשת מזוהה": best["רשת_מזוהה"] if best is not None else "",
            "מיקום במוסד חרדי": matched_haredi,
            "lat_anchor": float(best["lat"]) if anchor and best is not None else np.nan,
            "lon_anchor": float(best["lon"]) if anchor and best is not None else np.nan,
        })
    matches = pd.DataFrame(match_rows)
    sites = sites.merge(matches, on="מזהה_אתר", how="left")

    # Geographic density of Haredi institutions around reliably anchored polling sites.
    inst_geo = mos[mos["קואורדינטה_לצפיפות"]].copy().reset_index(drop=True)
    tree = BallTree(np.deg2rad(inst_geo[["lat", "lon"]].to_numpy()), metric="haversine")
    earth_km = 6371.0088
    density_rows: list[dict[str, Any]] = []
    for _, site in sites.iterrows():
        row: dict[str, Any] = {"מזהה_אתר": site["מזהה_אתר"]}
        for radius in [250, 500, 1000]:
            row[f"מספר מוסדות חרדיים {radius}מ"] = 0
            row[f"תלמידים חרדיים שקולים {radius}מ"] = 0.0
            row[f"שיעור מוסדות חרדיים {radius}מ"] = 0.0
        if bool(site.get("עוגן קואורדינטות", False)) and pd.notna(site.get("lat_anchor")):
            point = np.deg2rad([[float(site["lat_anchor"]), float(site["lon_anchor"])]] )
            inds, dists = tree.query_radius(point, r=1.0 / earth_km, return_distance=True, sort_results=True)
            idxs = inds[0]
            meters = dists[0] * earth_km * 1000
            nearby = inst_geo.iloc[idxs].copy()
            nearby["distance_m"] = meters
            for radius in [250, 500, 1000]:
                sub = nearby[nearby["distance_m"] <= radius]
                h = sub[sub["חרדי"]]
                h_eq = float(h["תלמידים_שקולים"].sum())
                all_eq = float(sub["תלמידים_שקולים"].sum())
                row[f"מספר מוסדות חרדיים {radius}מ"] = int(len(h))
                row[f"תלמידים חרדיים שקולים {radius}מ"] = round(h_eq, 2)
                row[f"שיעור מוסדות חרדיים {radius}מ"] = round(h_eq / all_eq, 4) if all_eq > 0 else 0.0
        density_rows.append(row)
    densities = pd.DataFrame(density_rows)
    sites = sites.merge(densities, on="מזהה_אתר", how="left")

    sites["צפיפות חרדית חזקה"] = (
        ((sites["מספר מוסדות חרדיים 250מ"] >= 1) & (sites["שיעור מוסדות חרדיים 250מ"] >= 0.65) & (sites["תלמידים חרדיים שקולים 250מ"] >= 60))
        | ((sites["מספר מוסדות חרדיים 500מ"] >= 2) & (sites["שיעור מוסדות חרדיים 500מ"] >= 0.50) & (sites["תלמידים חרדיים שקולים 500מ"] >= 120))
    )
    sites["צפיפות חרדית בינונית"] = (~sites["צפיפות חרדית חזקה"]) & (
        (sites["מספר מוסדות חרדיים 1000מ"] >= 2)
        & (sites["שיעור מוסדות חרדיים 1000מ"] >= 0.30)
        & (sites["תלמידים חרדיים שקולים 1000מ"] >= 120)
    )
    sites["אות מוסדי חזק"] = sites["מיקום במוסד חרדי"] | sites["צפיפות חרדית חזקה"]
    sites["אות מוסדי בינוני"] = (~sites["אות מוסדי חזק"]) & sites["צפיפות חרדית בינונית"]

    regular["מזהה_אתר"] = regular["סמל ישוב"].astype(int).astype(str) + "|" + regular["מיקום קלפי"].astype(str)
    regular = regular.merge(
        sites.drop(columns=["סמל ישוב", "שם ישוב", "עיר_נורמלית", "מיקום קלפי"], errors="ignore"),
        on="מזהה_אתר",
        how="left",
    )
    for c in ["מיקום במוסד חרדי", "צפיפות חרדית חזקה", "צפיפות חרדית בינונית", "אות מוסדי חזק", "אות מוסדי בינוני", "עוגן קואורדינטות", "התאמה התקבלה"]:
        regular[c] = regular[c].fillna(False).astype(bool)

    # Direct observed counts.
    total_shas = int(ballots["שס"].sum())
    total_g = int(ballots["ג"].sum())
    external_shas = int(ballots.loc[ballots["חיצוניות"], "שס"].sum())
    regular_shas = int(regular["שס"].sum())
    direct_rows: list[dict[str, Any]] = []
    for threshold in [0.20, 0.15, 0.10, 0.05]:
        mask = regular["שיעור ג"] >= threshold
        direct_rows.append({
            "הגדרה": f"יהדות התורה לפחות {threshold:.0%}",
            "מספר קלפיות": int(mask.sum()),
            "קולות שס": int(regular.loc[mask, "שס"].sum()),
            "אחוז מכלל קולות שס": round(100 * regular.loc[mask, "שס"].sum() / total_shas, 2),
        })
    for label, mask in [
        ("יישובים חרדיים מובהקים", regular["יישוב חרדי מובהק"]),
        ("קלפי בתוך מוסד חרדי מותאם", regular["מיקום במוסד חרדי"]),
        ("אות מוסדי חזק", regular["אות מוסדי חזק"]),
        ("אות מוסדי בינוני או חזק", regular["אות מוסדי חזק"] | regular["אות מוסדי בינוני"]),
        ("יהדות התורה 10%+ או אות מוסדי חזק", (regular["שיעור ג"] >= 0.10) | regular["אות מוסדי חזק"]),
    ]:
        direct_rows.append({
            "הגדרה": label,
            "מספר קלפיות": int(mask.sum()),
            "קולות שס": int(regular.loc[mask, "שס"].sum()),
            "אחוז מכלל קולות שס": round(100 * regular.loc[mask, "שס"].sum() / total_shas, 2),
        })
    direct = pd.DataFrame(direct_rows)

    low_g = regular["שיעור ג"] < 0.10
    catch_strong = low_g & regular["אות מוסדי חזק"]
    catch_medium = low_g & (~regular["אות מוסדי חזק"]) & regular["אות מוסדי בינוני"]
    catch_direct_host = low_g & regular["מיקום במוסד חרדי"]

    # Transparent sensitivity scenarios. These are area-probability weights, not individual identifications.
    def scenario_weights(df: pd.DataFrame, scenario: str) -> np.ndarray:
        g = df["שיעור ג"].to_numpy()
        core = df["יישוב חרדי מובהק"].to_numpy(bool)
        host = df["מיקום במוסד חרדי"].to_numpy(bool)
        strong = df["צפיפות חרדית חזקה"].to_numpy(bool)
        medium = df["צפיפות חרדית בינונית"].to_numpy(bool)
        w = np.zeros(len(df), dtype=float)
        if scenario == "שמרני":
            w = np.where(core, 0.98, w)
            w = np.where((~core) & (g >= 0.20), 0.92, w)
            w = np.where((w == 0) & (g >= 0.10), 0.62, w)
            w = np.where((w == 0) & host, 0.58, w)
            w = np.where((w == 0) & strong, 0.42, w)
            w = np.where((w == 0) & medium & (g >= 0.02), 0.18, w)
        elif scenario == "מרכזי":
            w = np.where(core, 0.98, w)
            w = np.where((~core) & (g >= 0.20), 0.95, w)
            w = np.where((w == 0) & (g >= 0.10), 0.78, w)
            w = np.where((w == 0) & host, 0.72, w)
            w = np.where((w == 0) & strong, 0.56, w)
            w = np.where((w == 0) & medium, 0.32, w)
            w = np.where((w == 0) & (g >= 0.05), 0.38, w)
            w = np.where((w == 0) & (g >= 0.02), 0.14, w)
        elif scenario == "רחב":
            w = np.where(core, 0.99, w)
            w = np.where((~core) & (g >= 0.15), 0.97, w)
            w = np.where((w == 0) & (g >= 0.10), 0.90, w)
            w = np.where((w == 0) & host, 0.86, w)
            w = np.where((w == 0) & strong, 0.72, w)
            w = np.where((w == 0) & medium, 0.50, w)
            w = np.where((w == 0) & (g >= 0.05), 0.58, w)
            w = np.where((w == 0) & (g >= 0.02), 0.28, w)
            w = np.where((w == 0) & (g >= 0.005), 0.06, w)
        else:
            raise ValueError(scenario)
        return w

    scenario_rows: list[dict[str, Any]] = []
    external_weights = {"שמרני": 0.10, "מרכזי": 0.17, "רחב": 0.27}
    for scenario in ["שמרני", "מרכזי", "רחב"]:
        w = scenario_weights(regular, scenario)
        regular_est = float(np.sum(regular["שס"].to_numpy() * w))
        external_est = external_shas * external_weights[scenario]
        total_est = regular_est + external_est
        # Counterfactual model with institutional indicators removed.
        shadow = regular.copy()
        shadow["מיקום במוסד חרדי"] = False
        shadow["צפיפות חרדית חזקה"] = False
        shadow["צפיפות חרדית בינונית"] = False
        base_w = scenario_weights(shadow, scenario)
        base_regular_est = float(np.sum(shadow["שס"].to_numpy() * base_w))
        scenario_rows.append({
            "תרחיש": scenario,
            "אומדן חרדים בקלפיות רגילות": int(round(regular_est)),
            "אומדן חרדים במעטפות חיצוניות": int(round(external_est)),
            "אומדן חרדים כולל": int(round(total_est)),
            "שיעור חרדים מכלל מצביעי שס": round(100 * total_est / total_shas, 2),
            "תוספת בזכות שכבת המוסדות": int(round(regular_est - base_regular_est)),
            "שיעור שיוחס לחרדים במעטפות": external_weights[scenario],
        })
    scenarios = pd.DataFrame(scenario_rows)

    # Diagnostics and validation proxies.
    core = regular[regular["יישוב חרדי מובהק"]]
    noncore = regular[~regular["יישוב חרדי מובהק"]]
    clear_non_haredi = noncore[(noncore["שיעור ג"] < 0.01) & (noncore["שיעור חרדי משולב"] < 0.12)]
    validation = pd.DataFrame([
        {"מדד": "קלפיות רגילות", "ערך": len(regular)},
        {"מדד": "קלפיות עם שם מיקום", "ערך": int((regular["מיקום קלפי"] != "").sum())},
        {"מדד": "קלפיות עם התאמת מוסד שהתקבלה", "ערך": int(regular["התאמה התקבלה"].sum())},
        {"מדד": "קלפיות עם עוגן קואורדינטות אמין", "ערך": int(regular["עוגן קואורדינטות"].sum())},
        {"מדד": "רגישות אות מוסדי חזק ביישובים חרדיים — לפי קולות שס", "ערך": round(100 * core.loc[core["אות מוסדי חזק"], "שס"].sum() / max(1, core["שס"].sum()), 2)},
        {"מדד": "שיעור אות מוסדי חזק בפרוקסי לא־חרדי ברור — לפי קולות שס", "ערך": round(100 * clear_non_haredi.loc[clear_non_haredi["אות מוסדי חזק"], "שס"].sum() / max(1, clear_non_haredi["שס"].sum()), 2)},
        {"מדד": "קולות שס שנלכדו באות מוסדי חזק כאשר ג מתחת ל־10%", "ערך": int(regular.loc[catch_strong, "שס"].sum())},
        {"מדד": "מתוכם קלפי בתוך מוסד חרדי", "ערך": int(regular.loc[catch_direct_host, "שס"].sum())},
        {"מדד": "קולות שס שנלכדו באות מוסדי בינוני בלבד כאשר ג מתחת ל־10%", "ערך": int(regular.loc[catch_medium, "שס"].sum())},
    ])

    # Candidate ballots and site-level aggregation that changed classification.
    candidates = regular[catch_strong | catch_medium].copy()
    candidates["סוג לכידה"] = np.where(candidates["אות מוסדי חזק"], "אות מוסדי חזק", "אות מוסדי בינוני")
    candidate_cols = [
        "שם ישוב", "קלפי_טקסט", "מיקום קלפי", "כשרים", "שס", "ג", "שיעור שס", "שיעור ג",
        "שיעור חרדי משולב", "סוג לכידה", "מיקום במוסד חרדי", "צפיפות חרדית חזקה",
        "צפיפות חרדית בינונית", "שם מוסד מותאם", "פיקוח מוסד מותאם", "רשת מזוהה",
        "ציון התאמה", "מספר מוסדות חרדיים 500מ", "תלמידים חרדיים שקולים 500מ",
        "שיעור מוסדות חרדיים 500מ", "מספר מוסדות חרדיים 1000מ",
        "תלמידים חרדיים שקולים 1000מ", "שיעור מוסדות חרדיים 1000מ",
    ]
    candidates = candidates[candidate_cols].sort_values(["שס", "שם ישוב"], ascending=[False, True])

    by_city = regular.assign(
        נתפס_חזק=catch_strong,
        נתפס_בינוני=catch_medium,
    ).groupby("שם ישוב", as_index=False).agg(
        קולות_שס=("שס", "sum"),
        קולות_שס_בג_10_ומעלה=("שס", lambda x: int(x[regular.loc[x.index, "שיעור ג"] >= 0.10].sum())),
        קולות_שס_שנתפסו_מוסדית_חזק=("שס", lambda x: int(x[regular.loc[x.index, "נתפס_חזק"]].sum())),
        קולות_שס_שנתפסו_מוסדית_בינוני=("שס", lambda x: int(x[regular.loc[x.index, "נתפס_בינוני"]].sum())),
        מספר_קלפיות=("שס", "size"),
    )
    by_city["תוספת מוסדית"] = by_city["קולות_שס_שנתפסו_מוסדית_חזק"] + by_city["קולות_שס_שנתפסו_מוסדית_בינוני"]
    by_city = by_city.sort_values("תוספת מוסדית", ascending=False)

    summary = {
        "total_shas": total_shas,
        "total_utj": total_g,
        "regular_shas": regular_shas,
        "external_shas": external_shas,
        "official_ballot_rows": int(len(ballots)),
        "regular_ballot_rows": int(len(regular)),
        "institution_year": target_year,
        "institution_rows": int(len(mos)),
        "haredi_institution_rows": int(mos["חרדי"].sum()),
        "matched_regular_ballots": int(regular["התאמה התקבלה"].sum()),
        "anchored_regular_ballots": int(regular["עוגן קואורדינטות"].sum()),
        "strong_catch_low_utj_shas": int(regular.loc[catch_strong, "שס"].sum()),
        "medium_catch_low_utj_shas": int(regular.loc[catch_medium, "שס"].sum()),
        "direct_host_low_utj_shas": int(regular.loc[catch_direct_host, "שס"].sum()),
        "scenarios": scenario_rows,
    }
    audit["summary"] = summary

    # Export full inspectable data.
    regular_export_cols = [
        "שם ישוב", "סמל ישוב", "קלפי_טקסט", "מיקום קלפי", "כשרים", "שס", "ג", "שיעור שס", "שיעור ג",
        "שיעור חרדי משולב", "יישוב חרדי מובהק", "מיקום במוסד חרדי", "אות מוסדי חזק",
        "אות מוסדי בינוני", "התאמה התקבלה", "ציון התאמה", "שם מוסד מותאם", "פיקוח מוסד מותאם",
        "רשת מזוהה", "עוגן קואורדינטות", "מספר מוסדות חרדיים 250מ", "תלמידים חרדיים שקולים 250מ",
        "שיעור מוסדות חרדיים 250מ", "מספר מוסדות חרדיים 500מ", "תלמידים חרדיים שקולים 500מ",
        "שיעור מוסדות חרדיים 500מ", "מספר מוסדות חרדיים 1000מ", "תלמידים חרדיים שקולים 1000מ",
        "שיעור מוסדות חרדיים 1000מ",
    ]
    regular_export = regular[regular_export_cols].copy()
    for c in ["שיעור שס", "שיעור ג", "שיעור חרדי משולב", "שיעור מוסדות חרדיים 250מ", "שיעור מוסדות חרדיים 500מ", "שיעור מוסדות חרדיים 1000מ"]:
        regular_export[c] = (regular_export[c] * 100).round(2)
    regular_export.to_csv(OUT / "ballot_classification.csv", index=False, encoding="utf-8-sig")
    candidates.to_csv(OUT / "institution_catch_candidates.csv", index=False, encoding="utf-8-sig")
    by_city.to_csv(OUT / "institution_uplift_by_city.csv", index=False, encoding="utf-8-sig")
    sites.to_csv(OUT / "polling_site_matches.csv", index=False, encoding="utf-8-sig")
    direct.to_csv(OUT / "direct_counts.csv", index=False, encoding="utf-8-sig")
    scenarios.to_csv(OUT / "scenarios.csv", index=False, encoding="utf-8-sig")
    validation.to_csv(OUT / "validation.csv", index=False, encoding="utf-8-sig")
    (OUT / "audit.json").write_text(json.dumps(audit, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    # Excel workbook.
    wb = Workbook()
    wb.remove(wb.active)
    summary_df = pd.DataFrame([
        ["כלל קולות שס", total_shas],
        ["קולות שס בקלפיות רגילות", regular_shas],
        ["קולות שס במעטפות חיצוניות", external_shas],
        ["שנת נתוני המוסדות", target_year],
        ["מוסדות חינוך בשנת החיתוך", len(mos)],
        ["מוסדות שסווגו חרדיים", int(mos["חרדי"].sum())],
        ["קלפיות רגילות עם התאמת מוסד", int(regular["התאמה התקבלה"].sum())],
        ["קלפיות רגילות עם עוגן קואורדינטות", int(regular["עוגן קואורדינטות"].sum())],
        ["קולות שס שנוספו באות מוסדי חזק מתחת ל־10% ג", int(regular.loc[catch_strong, "שס"].sum())],
        ["קולות שס שנוספו באות מוסדי בינוני מתחת ל־10% ג", int(regular.loc[catch_medium, "שס"].sum())],
    ], columns=["מדד", "ערך"])
    add_sheet(wb, "תקציר", summary_df, {"מדד": 55, "ערך": 18})
    add_sheet(wb, "תרחישים", scenarios)
    add_sheet(wb, "ספירות ישירות", direct)
    add_sheet(wb, "בדיקות איכות", validation, {"מדד": 70, "ערך": 20})
    add_sheet(wb, "תוספת לפי עיר", by_city)
    add_sheet(wb, "קלפיות שנלכדו", candidates)
    add_sheet(wb, "כל הקלפיות", regular_export)
    site_export = sites.drop(columns=["lat_anchor", "lon_anchor"], errors="ignore")
    add_sheet(wb, "התאמות אתרי הצבעה", site_export)
    wb.save(OUT / "shas_2022_haredi_institution_analysis.xlsx")

    # Human-readable report.
    central = scenarios.loc[scenarios["תרחיש"] == "מרכזי"].iloc[0]
    strict = scenarios.loc[scenarios["תרחיש"] == "שמרני"].iloc[0]
    broad = scenarios.loc[scenarios["תרחיש"] == "רחב"].iloc[0]
    report = f"""# מחקר מצביעי ש״ס 2022 — שכבת מוסדות חינוך חרדיים

## תוצאה ראשונית

בתרחיש המרכזי נאמדו **{central['אומדן חרדים כולל']:,}** מצביעי ש״ס חרדים, שהם **{central['שיעור חרדים מכלל מצביעי שס']:.1f}%** מכלל {total_shas:,} קולות ש״ס.
טווח הרגישות הוא **{strict['אומדן חרדים כולל']:,}–{broad['אומדן חרדים כולל']:,}**, כלומר **{strict['שיעור חרדים מכלל מצביעי שס']:.1f}%–{broad['שיעור חרדים מכלל מצביעי שס']:.1f}%**.

שכבת מוסדות החינוך הוסיפה בתרחיש המרכזי **{central['תוספת בזכות שכבת המוסדות']:,}** קולות לאומדן החרדי, מעבר למודל המבוסס על דפוס ההצבעה ליהדות התורה בלבד.

## מה נמצא ישירות בנתונים

- בקובץ הרשמי: {len(ballots):,} רשומות קלפי ו־{total_shas:,} קולות ש״ס.
- במעטפות החיצוניות: {external_shas:,} קולות ש״ס, שאין להם שיוך גיאוגרפי למקום מגורי הבוחר.
- מתחת לסף של 10% ליהדות התורה, שכבת המוסדות זיהתה {int(regular.loc[catch_strong, 'שס'].sum()):,} קולות ש״ס באות מוסדי חזק ועוד {int(regular.loc[catch_medium, 'שס'].sum()):,} באות בינוני.
- {int(regular.loc[catch_direct_host, 'שס'].sum()):,} מן הקולות שמתחת לסף 10% נמצאו בקלפיות שהותאמו ישירות למוסד חרדי בפיקוח משרד החינוך.

## פירוש

המספר אינו זיהוי אישי של בוחרים אלא אומדן אקולוגי ברמת הקלפי. שכבת המוסדות מועילה בעיקר בזיהוי מוקדים ספרדיים־חרדיים שבהם שיעור ההצבעה ליהדות התורה נמוך, אך אתר ההצבעה או סביבתו החינוכית הם חרדיים מובהקים.

## מקורות ושיטה

1. תוצאות האמת הרשמיות של הכנסת ה־25 לפי קלפיות.
2. מיקום אתר ההצבעה לכל מספר קלפי.
3. מאפייני מוסדות החינוך לשנת {target_year}, כולל פיקוח, מספר תלמידים וסוג מוסד.
4. קואורדינטות שערי הכניסה למוסדות החינוך; קואורדינטות ברמת דיוק נמוכה הוצאו מחישובי הצפיפות.
5. התאמת שמות בתוך אותו יישוב, ולאחר מכן חישוב מוסדות חרדיים ברדיוסים של 250, 500 ו־1,000 מטר.
6. שלושה תרחישי משקל מפורשים כדי למנוע הצגת דיוק כוזב.
"""
    (OUT / "report.md").write_text(report, encoding="utf-8")

    print("RESULT_JSON_START")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print("RESULT_JSON_END")
    print("TOP_CITIES_START")
    print(by_city.head(25).to_json(orient="records", force_ascii=False))
    print("TOP_CITIES_END")
    print("VALIDATION_START")
    print(validation.to_json(orient="records", force_ascii=False))
    print("VALIDATION_END")


if __name__ == "__main__":
    main()
